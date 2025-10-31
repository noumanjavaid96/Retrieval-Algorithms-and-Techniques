import openpyxl

def parse_xlsx(file_path):
    workbook = openpyxl.load_workbook(file_path)
    text = ""
    for sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    text += str(cell.value) + " "
            text += "\n"
    return text
